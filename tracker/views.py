from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Sum
from django.core.mail import send_mail
from django.http import JsonResponse
from django.utils import timezone
from datetime import date, timedelta
from decimal import Decimal
import calendar

from .models import Profile, Transaction, Budget, Reminder, Category
from django.contrib.auth.forms import UserCreationForm

# ----------------- Auth Views -----------------

def signup_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            auth_login(request, user)
            # Create sample data for the user on signup to make it look great instantly
            seed_user_demo_data(user)
            messages.success(request, f"Welcome to Leoxur tracker, {user.username}!")
            return redirect('dashboard')
    else:
        form = UserCreationForm()
    return render(request, 'tracker/signup.html', {'form': form})

# ----------------- Dashboard & Analytics -----------------

@login_required
def dashboard_view(request):
    today = date.today()
    user = request.user
    profile, _ = Profile.objects.get_or_create(user=user)
    
    # Asynchronously fetch incoming email transactions in a background thread if auto-fetch is enabled
    if profile.auto_fetch_emails:
        import threading
        from django.core.management import call_command
        from django.db import connections
        
        def run_background_sync():
            try:
                call_command('fetch_emails')
            except Exception as e:
                print(f"Background email sync failed: {e}")
            finally:
                connections.close_all()

        try:
            threading.Thread(target=run_background_sync).start()
        except Exception as e:
            print(f"Background email fetch trigger failed: {e}")
    
    # Dynamic Category Seeding Fallback (for legacy or existing users)
    if not Category.objects.filter(user=user).exists():
        defaults = [
            ('Salary', True),
            ('Freelance', True),
            ('Investment', True),
            ('Food', False),
            ('Rent', False),
            ('Utilities', False),
            ('Entertainment', False),
            ('Travel', False),
            ('Other', False),
        ]
        for name, is_income in defaults:
            Category.objects.get_or_create(user=user, name=name, is_income=is_income)
            
    user_income_cats = Category.objects.filter(user=user, is_income=True)
    user_expense_cats = Category.objects.filter(user=user, is_income=False)
    
    income_categories = [c.name for c in user_income_cats]
    expense_categories = [c.name for c in user_expense_cats]
    
    # Filter bounds for current month
    start_of_month = date(today.year, today.month, 1)
    _, last_day = calendar.monthrange(today.year, today.month)
    end_of_month = date(today.year, today.month, last_day)

    # Current Month Transactions
    month_txs = Transaction.objects.filter(user=user, date__range=(start_of_month, end_of_month))
    
    # Financial Aggregations
    total_income = month_txs.filter(transaction_type='IN').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    total_expense = month_txs.filter(transaction_type='OUT').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    balance = total_income - total_expense
    
    # Category Expenses for current month
    cat_expenses = {}
    for tx in month_txs.filter(transaction_type='OUT'):
        cat_expenses[tx.category] = cat_expenses.get(tx.category, Decimal('0.00')) + tx.amount

    # Budget Progress Calculations
    budgets = Budget.objects.filter(user=user, year=today.year, month=today.month)
    budget_data = []
    alerts = []
    
    for category in expense_categories:
        budget_obj = budgets.filter(category=category).first()
        budget_amount = budget_obj.amount if budget_obj else Decimal('0.00')
        spent = cat_expenses.get(category, Decimal('0.00'))
        
        pct = 0
        if budget_amount > 0:
            pct = int((spent / budget_amount) * 100)
            if pct >= 100:
                alerts.append({
                    'type': 'danger',
                    'message': f"CRITICAL: You have exceeded your monthly budget for {category}! Spent: {profile.currency}{spent:.2f} / Budget: {profile.currency}{budget_amount:.2f} ({pct}%)"
                })
            elif pct >= 80:
                alerts.append({
                    'type': 'warning',
                    'message': f"WARNING: You've reached {pct}% of your monthly budget for {category}. Remaining: {profile.currency}{budget_amount - spent:.2f}"
                })
        
        remaining = budget_amount - spent
        overage = spent - budget_amount
        budget_data.append({
            'category': category,
            'budget': budget_amount,
            'spent': spent,
            'percentage': min(pct, 100) if budget_amount > 0 else 0,
            'raw_percentage': pct,
            'has_budget': budget_amount > 0,
            'remaining': remaining,
            'overage': overage,
        })

    # Overall Total Budget
    overall_budget_obj = budgets.filter(category='Total').first()
    overall_budget = overall_budget_obj.amount if overall_budget_obj else Decimal('0.00')
    overall_pct = 0
    if overall_budget > 0:
        overall_pct = int((total_expense / overall_budget) * 100)
        if overall_pct >= 100:
            alerts.append({
                'type': 'danger',
                'message': f"CRITICAL: Total expenses have breached your overall budget! Spent: {profile.currency}{total_expense:.2f} / Budget: {profile.currency}{overall_budget:.2f} ({overall_pct}%)"
            })
        elif overall_pct >= 80:
            alerts.append({
                'type': 'warning',
                'message': f"WARNING: Total expenses at {overall_pct}% of overall budget. Remaining: {profile.currency}{overall_budget - total_expense:.2f}"
            })

    # Smart Saving Recommendations
    recommendations = generate_recommendations(user, cat_expenses, total_income)

    # Reminders
    reminders = Reminder.objects.filter(user=user, is_paid=False)
    upcoming_reminders = reminders.filter(due_date__gte=today)
    overdue_reminders = reminders.filter(due_date__lt=today)
    
    recent_transactions = Transaction.objects.filter(user=user).order_by('-date', '-created_at')[:500]

    # Calculate gamified Financial Health Score
    health_score = 100
    if total_income > 0:
        outflow_ratio = total_expense / total_income
        if outflow_ratio > 1.0:
            health_score -= 30
        elif outflow_ratio > 0.9:
            health_score -= 15
        elif outflow_ratio > 0.8:
            health_score -= 5
    else:
        if total_expense > 0:
            health_score -= 30

    breached_count = 0
    for category in expense_categories:
        b_obj = budgets.filter(category=category).first()
        if b_obj and b_obj.amount > 0:
            spent = cat_expenses.get(category, Decimal('0.00'))
            if spent > b_obj.amount:
                breached_count += 1
    health_score -= min(breached_count * 15, 35)

    overdue_count = overdue_reminders.count()
    health_score -= min(overdue_count * 10, 30)
    health_score = max(10, health_score)
    health_offset = float(175.9 * (1.0 - health_score / 100.0))

    context = {
        'profile': profile,
        'total_income': total_income,
        'total_expense': total_expense,
        'balance': balance,
        'budget_data': budget_data,
        'overall_budget': overall_budget,
        'overall_percentage': min(overall_pct, 100) if overall_budget > 0 else 0,
        'overall_raw_percentage': overall_pct,
        'overall_remaining': overall_budget - total_expense,
        'alerts': alerts,
        'recommendations': recommendations,
        'upcoming_reminders': upcoming_reminders,
        'overdue_reminders': overdue_reminders,
        'recent_transactions': recent_transactions,
        'user_income_cats': user_income_cats,
        'user_expense_cats': user_expense_cats,
        'expense_categories': expense_categories,
        'today': today,
        'health_score': health_score,
        'health_offset': health_offset,
    }
    return render(request, 'tracker/dashboard.html', context)


def generate_recommendations(user, current_expenses, current_income):
    recommendations = []
    currency = user.profile.currency
    total_expenses = sum(current_expenses.values())

    # 1. Total expenses vs Income
    if total_expenses > current_income and current_income > 0:
        pct = int((total_expenses / current_income) * 100)
        recommendations.append(
            f"Alert: Outflows are {pct}% of your income. Consider suspending non-essential subscriptions and entertainment spending."
        )

    # 2. Category spending comparison to last month
    today = date.today()
    first_of_this_month = date(today.year, today.month, 1)
    last_month_date = first_of_this_month - timedelta(days=1)
    
    last_month_transactions = Transaction.objects.filter(
        user=user,
        transaction_type='OUT',
        date__year=last_month_date.year,
        date__month=last_month_date.month
    )
    
    last_month_expenses = {}
    for tx in last_month_transactions:
        last_month_expenses[tx.category] = last_month_expenses.get(tx.category, Decimal('0.00')) + tx.amount

    for category, current_amount in current_expenses.items():
        prev_amount = last_month_expenses.get(category, Decimal('0.00'))
        if prev_amount > 0 and current_amount > prev_amount:
            increase = current_amount - prev_amount
            increase_pct = int((increase / prev_amount) * 100)
            if increase_pct >= 15:
                recommendations.append(
                    f"You spent {increase_pct}% more on {category} this month than last month. Cutting back could save you {currency}{increase:.2f}."
                )

    # 3. Large concentration of expenses
    if total_expenses > 0:
        for category, amount in current_expenses.items():
            pct = (amount / total_expenses) * 100
            if pct > 30 and category not in ['Rent', 'Other']:
                recommendations.append(
                    f"{category} accounts for {int(pct)}% of all outflows. Setting a tighter budget here will accelerate your savings."
                )

    # 4. No Budgets set check
    user_budgets = Budget.objects.filter(user=user, year=today.year, month=today.month)
    if not user_budgets.exists():
        recommendations.append(
            "You haven't set any monthly budgets! Users with target budgets save an average of 18% more monthly."
        )

    if not recommendations:
        recommendations.append(
            "Tip: Review your active category budget targets to optimize monthly surplus outflows."
        )

    return recommendations[:3]

# ----------------- AJAX/JSON Analytics Data -----------------

@login_required
def analytics_data_api(request):
    user = request.user
    today = date.today()
    
    # 1. 6-Month Trends: Income vs Expense
    trends = []
    # Generate last 6 months list
    for i in range(5, -1, -1):
        # Subtract months
        year = today.year
        month = today.month - i
        if month <= 0:
            month += 12
            year -= 1
        
        _, last_day = calendar.monthrange(year, month)
        start = date(year, month, 1)
        end = date(year, month, last_day)
        
        txs = Transaction.objects.filter(user=user, date__range=(start, end))
        inc = txs.filter(transaction_type='IN').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        exp = txs.filter(transaction_type='OUT').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        
        trends.append({
            'month': start.strftime('%b %Y'),
            'income': float(inc),
            'expense': float(exp),
        })

    # 2. Current Month Category Expense Breakdown
    start_of_month = date(today.year, today.month, 1)
    _, last_day = calendar.monthrange(today.year, today.month)
    end_of_month = date(today.year, today.month, last_day)
    
    month_exp_txs = Transaction.objects.filter(
        user=user, 
        transaction_type='OUT', 
        date__range=(start_of_month, end_of_month)
    )
    
    category_totals = {}
    for tx in month_exp_txs:
        category_totals[tx.category] = category_totals.get(tx.category, 0.0) + float(tx.amount)
        
    categories = list(category_totals.keys())
    values = list(category_totals.values())
    
    # 3. Weekly Outflow Distribution (Week 1: 1-7, Week 2: 8-14, Week 3: 15-21, Week 4: 22-end)
    weekly_sums = [0.0, 0.0, 0.0, 0.0]
    for tx in month_exp_txs:
        day = tx.date.day
        if day <= 7:
            weekly_sums[0] += float(tx.amount)
        elif day <= 14:
            weekly_sums[1] += float(tx.amount)
        elif day <= 21:
            weekly_sums[2] += float(tx.amount)
        else:
            weekly_sums[3] += float(tx.amount)

    # 4. Budget vs Spent
    expense_categories = ['Food', 'Rent', 'Utilities', 'Entertainment', 'Travel', 'Other']
    budgets = Budget.objects.filter(user=user, year=today.year, month=today.month)
    
    comp_labels = []
    comp_budgets = []
    comp_spent = []
    
    for cat in expense_categories:
        b_obj = budgets.filter(category=cat).first()
        b_amount = float(b_obj.amount) if b_obj else 0.0
        s_amount = float(category_totals.get(cat, 0.0))
        if b_amount > 0 or s_amount > 0:
            comp_labels.append(cat)
            comp_budgets.append(b_amount)
            comp_spent.append(s_amount)

    # Return everything in single clean payload
    return JsonResponse({
        'trends': trends,
        'category_breakdown': {
            'labels': categories,
            'data': values
        },
        'budget_vs_spent': {
            'labels': comp_labels,
            'budgets': comp_budgets,
            'spent': comp_spent
        },
        'weekly_distribution': {
            'labels': ['Week 1', 'Week 2', 'Week 3', 'Week 4'],
            'data': weekly_sums
        }
    })


@login_required
def savings_analyst_chat(request):
    if request.method == 'POST':
        import json
        import re
        from datetime import datetime, date
        import calendar
        from decimal import Decimal
        from django.db.models import Sum, Q
        from django.core.management import call_command
        import threading

        try:
            body = json.loads(request.body)
            query = body.get('query', '').strip()
        except Exception:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON body'}, status=400)
            
        if not query:
            return JsonResponse({'status': 'error', 'message': 'Query cannot be empty'}, status=400)
            
        user = request.user
        profile = user.profile
        currency = profile.currency
        today = date.today()
        start_of_month = date(today.year, today.month, 1)
        _, last_day = calendar.monthrange(today.year, today.month)
        end_of_month = date(today.year, today.month, last_day)

        def parse_date(date_str):
            for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d'):
                try:
                    return datetime.strptime(date_str.strip(), fmt).date()
                except ValueError:
                    pass
            return today

        # ----------------- PARSE COMMANDS -----------------
        q = query.strip()
        q_lower = q.lower()

        # 1. Sync Emails Command
        if re.match(r'^(?:sync\s+emails|fetch\s+emails|sync\s+transactions)$', q_lower):
            from django.db import connections
            def run_background_sync():
                try:
                    call_command('fetch_emails')
                except Exception as e:
                    print(f"Background email sync failed: {e}")
                finally:
                    connections.close_all()
            try:
                threading.Thread(target=run_background_sync).start()
                response = (
                    f"<div class='space-y-2'>"
                    f"  <div class='flex items-center space-x-1.5'>"
                    f"    <span class='px-2 py-0.5 text-[9px] font-bold bg-neon-cyan/10 text-neon-cyan rounded border border-neon-cyan/20 uppercase'>System Link</span>"
                    f"    <strong class='text-xs text-slate-800 dark:text-white'>Inbox Sync Initiated</strong>"
                    f"  </div>"
                    f"  <p class='text-xs text-slate-600 dark:text-slate-300'>I've triggered a secure fetch sequence to scan your configured email node in the background. Fresh incoming transactions will populate your dashboard momentarily.</p>"
                    f"</div>"
                )
            except Exception as e:
                response = f"Failed to initiate background sync: {e}"
            return JsonResponse({'status': 'success', 'response': response})

        # 2. Add Transaction Command
        tx_match = re.match(
            r"^(?:add|log|record)\s+(expense|outflow|spend|income|inflow)\s+(?:of\s+)?([\d.,]+)\s+(?:for|on|in|to)?\s*([a-zA-Z0-9_ -]+?)(?:\s+(?:with\s+)?(?:description|desc)\s+(.+))?$",
            q,
            re.IGNORECASE
        )
        if tx_match:
            tx_type_str = tx_match.group(1).lower()
            amount_val = Decimal(tx_match.group(2).replace(',', ''))
            category_name = tx_match.group(3).strip().title()
            description = tx_match.group(4).strip() if tx_match.group(4) else ""

            tx_type = 'OUT' if tx_type_str in ('expense', 'outflow', 'spend') else 'IN'
            
            # Ensure category exists
            Category.objects.get_or_create(user=user, name=category_name, is_income=(tx_type == 'IN'))

            # Create Transaction
            tx = Transaction.objects.create(
                user=user,
                amount=amount_val,
                transaction_type=tx_type,
                category=category_name,
                date=today,
                description=description
            )
            
            if tx_type == 'OUT':
                check_and_send_budget_alerts(user, category_name, today)
                
            type_label = "Outflow logged" if tx_type == 'OUT' else "Inflow logged"
            color_class = "text-neon-rose bg-neon-rose/10 border-neon-rose/20" if tx_type == 'OUT' else "text-neon-green bg-neon-green/10 border-neon-green/20"
            
            response = (
                f"<div class='space-y-2'>"
                f"  <div class='flex items-center space-x-1.5'>"
                f"    <span class='px-2 py-0.5 text-[9px] font-bold rounded border uppercase {color_class}'>{type_label}</span>"
                f"    <strong class='text-xs text-slate-800 dark:text-white'>Operation Successful</strong>"
                f"  </div>"
                f"  <p class='text-xs text-slate-600 dark:text-slate-300'>Successfully added transaction to the financial ledger:</p>"
                f"  <ul class='list-disc pl-4 text-xs text-slate-600 dark:text-slate-300 space-y-0.5'>"
                f"    <li><strong>Amount:</strong> {currency}{tx.amount:.2f}</li>"
                f"    <li><strong>Category:</strong> {tx.category}</li>"
                f"    <li><strong>Date:</strong> {tx.date.strftime('%b %d, %Y')}</li>"
                f"    <li><strong>Description:</strong> {tx.description or 'None'}</li>"
                f"  </ul>"
                f"</div>"
            )
            return JsonResponse({'status': 'success', 'response': response})

        # 3. Set Budget Command
        budget_match = re.match(
            r"^(?:set|create|update)\s+(?:budget|limit)\s+(?:for|on)?\s*([a-zA-Z0-9_ -]+?)\s+(?:to|of)?\s*([\d.,]+)$",
            q,
            re.IGNORECASE
        )
        if not budget_match:
            budget_match = re.match(
                r"^budget\s+([a-zA-Z0-9_ -]+?)\s+([\d.,]+)$",
                q,
                re.IGNORECASE
            )
        if budget_match:
            category_name = budget_match.group(1).strip().title()
            amount_val = Decimal(budget_match.group(2).replace(',', ''))

            # Ensure category exists as expense
            Category.objects.get_or_create(user=user, name=category_name, is_income=False)

            budget, created = Budget.objects.update_or_create(
                user=user,
                category=category_name,
                period='MONTHLY',
                month=today.month,
                year=today.year,
                defaults={'amount': amount_val}
            )

            status_label = "Budget configured" if created else "Budget updated"
            response = (
                f"<div class='space-y-2'>"
                f"  <div class='flex items-center space-x-1.5'>"
                f"    <span class='px-2 py-0.5 text-[9px] font-bold bg-neon-purple/10 text-neon-purple rounded border border-neon-purple/20 uppercase'>{status_label}</span>"
                f"    <strong class='text-xs text-slate-800 dark:text-white'>Target Established</strong>"
                f"  </div>"
                f"  <p class='text-xs text-slate-600 dark:text-slate-300'>Configured category limit successfully:</p>"
                f"  <ul class='list-disc pl-4 text-xs text-slate-600 dark:text-slate-300 space-y-0.5'>"
                f"    <li><strong>Category:</strong> {budget.category}</li>"
                f"    <li><strong>Limit:</strong> {currency}{budget.amount:.2f}</li>"
                f"    <li><strong>Period:</strong> Monthly ({today.strftime('%B %Y')})</li>"
                f"  </ul>"
                f"</div>"
            )
            return JsonResponse({'status': 'success', 'response': response})

        # 4. Add Bill Reminder Command
        reminder_match = re.match(
            r"^(?:remind\s+me|add\s+reminder)\s+(?:to\s+pay\s+)?(.+?)\s+(?:of|amount)?\s*([\d.,]+)\s+(?:due\s+)?(?:on|by)?\s*([\d\-\/]+)$",
            q,
            re.IGNORECASE
        )
        if reminder_match:
            title = reminder_match.group(1).strip().title()
            amount_val = Decimal(reminder_match.group(2).replace(',', ''))
            date_str = reminder_match.group(3).strip()
            due_date = parse_date(date_str)

            reminder = Reminder.objects.create(
                user=user,
                title=title,
                amount=amount_val,
                due_date=due_date,
                is_recurring=False,
                is_paid=False
            )

            response = (
                f"<div class='space-y-2'>"
                f"  <div class='flex items-center space-x-1.5'>"
                f"    <span class='px-2 py-0.5 text-[9px] font-bold bg-neon-violet/10 text-neon-violet rounded border border-neon-violet/20 uppercase'>Reminder Scheduled</span>"
                f"    <strong class='text-xs text-slate-800 dark:text-white'>Outflow Alert Locked</strong>"
                f"  </div>"
                f"  <p class='text-xs text-slate-600 dark:text-slate-300'>Successfully added a new bill reminder:</p>"
                f"  <ul class='list-disc pl-4 text-xs text-slate-600 dark:text-slate-300 space-y-0.5'>"
                f"    <li><strong>Title:</strong> {reminder.title}</li>"
                f"    <li><strong>Amount:</strong> {currency}{reminder.amount:.2f}</li>"
                f"    <li><strong>Due Date:</strong> {reminder.due_date.strftime('%b %d, %Y')}</li>"
                f"  </ul>"
                f"</div>"
            )
            return JsonResponse({'status': 'success', 'response': response})

        # ----------------- PARSE QUERIES / ANALYSIS -----------------
        
        # Help Menu
        if q_lower in ('help', 'commands', 'menu', '?', 'what can you do'):
            response = (
                f"<div class='space-y-2 max-h-80 overflow-y-auto pr-1 text-xs text-slate-600 dark:text-slate-300'>"
                f"  <div class='flex items-center space-x-1.5 mb-1'>"
                f"    <span class='px-2 py-0.5 text-[9px] font-bold bg-neon-cyan/10 text-neon-cyan rounded border border-neon-cyan/20 uppercase'>Node Support</span>"
                f"    <strong class='text-xs text-slate-800 dark:text-white'>LEOXUR Analyst Commands</strong>"
                f"  </div>"
                f"  <p>I am your financial copilot. Here is a list of commands and queries you can run directly:</p>"
                f"  <div class='space-y-2.5 mt-2 pl-1'>"
                f"    <div>"
                f"      <strong class='text-neon-cyan'>1. Ledger Logging (Database Write)</strong>"
                f"      <p class='text-[10px] text-slate-400 dark:text-slate-500'>add expense/income [amount] [category] desc [details]</p>"
                f"      <code class='block bg-slate-100 dark:bg-space-950 p-1.5 rounded mt-1 border border-slate-200 dark:border-space-800 text-[10px]'>add expense 45 for Food desc lunch at office</code>"
                f"    </div>"
                f"    <div>"
                f"      <strong class='text-neon-cyan'>2. Budget Configuration</strong>"
                f"      <p class='text-[10px] text-slate-400 dark:text-slate-500'>set budget/limit [category] [amount]</p>"
                f"      <code class='block bg-slate-100 dark:bg-space-950 p-1.5 rounded mt-1 border border-slate-200 dark:border-space-800 text-[10px]'>set budget for Entertainment to 250</code>"
                f"    </div>"
                f"    <div>"
                f"      <strong class='text-neon-cyan'>3. Outflow Alerts (Bill Reminders)</strong>"
                f"      <p class='text-[10px] text-slate-400 dark:text-slate-500'>remind me to pay [title] [amount] on [YYYY-MM-DD]</p>"
                f"      <code class='block bg-slate-100 dark:bg-space-950 p-1.5 rounded mt-1 border border-slate-200 dark:border-space-800 text-[10px]'>remind me to pay internet bill 60 on 2026-07-25</code>"
                f"    </div>"
                f"    <div>"
                f"      <strong class='text-neon-cyan'>4. General Analysis & Actions</strong>"
                f"      <ul class='list-disc pl-4 space-y-1 text-[11px] mt-1'>"
                f"        <li><strong>Cashflow Status:</strong> Ask <em>'how is my cashflow'</em> or <em>'savings rate'</em></li>"
                f"        <li><strong>Check Budgets:</strong> Ask <em>'check budgets'</em> or <em>'are my budgets overspent?'</em></li>"
                f"        <li><strong>Search Ledger:</strong> Ask <em>'find Rent'</em> or <em>'search Food'</em></li>"
                f"        <li><strong>Savings Forecast:</strong> Ask <em>'predict savings'</em> or <em>'forecast'</em></li>"
                f"        <li><strong>Email Sync:</strong> Type <em>'sync emails'</em> or <em>'fetch transactions'</em></li>"
                f"      </ul>"
                f"    </div>"
                f"  </div>"
                f"</div>"
            )
            return JsonResponse({'status': 'success', 'response': response})

        # Gather general transaction aggregates
        month_txs = Transaction.objects.filter(user=user, date__range=(start_of_month, end_of_month))
        total_income = float(month_txs.filter(transaction_type='IN').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00'))
        total_expense = float(month_txs.filter(transaction_type='OUT').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00'))
        balance = total_income - total_expense
        
        category_totals = {}
        for tx in month_txs.filter(transaction_type='OUT'):
            category_totals[tx.category] = category_totals.get(tx.category, 0.0) + float(tx.amount)
            
        budgets = Budget.objects.filter(user=user, year=today.year, month=today.month)

        # Budget Check query
        if re.search(r'(?:budget|limit|overspent|spent)', q_lower):
            if budgets.exists():
                lines = []
                overspent_count = 0
                for b in budgets:
                    spent = category_totals.get(b.category, 0.0)
                    rem = float(b.amount) - spent
                    if rem < 0:
                        status_badge = "<span class='px-1.5 py-0.5 text-[9px] font-bold bg-neon-rose/10 text-neon-rose rounded border border-neon-rose/20 uppercase'>OVERSPENT</span>"
                        overspent_count += 1
                        val_str = f"<span class='text-neon-rose font-semibold'>-{currency}{abs(rem):.2f}</span>"
                    else:
                        status_badge = "<span class='px-1.5 py-0.5 text-[9px] font-bold bg-neon-green/10 text-neon-green rounded border border-neon-green/20 uppercase'>OK</span>"
                        val_str = f"<span class='text-slate-400 dark:text-slate-500'>rem. {currency}{rem:.2f}</span>"
                    
                    lines.append(
                        f"    <div class='flex justify-between items-center text-xs border-b border-slate-100 dark:border-space-850 pb-1.5'>"
                        f"      <div>"
                        f"        <strong class='text-slate-700 dark:text-slate-200'>{b.category}</strong>"
                        f"        <p class='text-[10px] text-slate-400 dark:text-slate-500'>Limit: {currency}{b.amount:.2f} | Spent: {currency}{spent:.2f}</p>"
                        f"      </div>"
                        f"      <div class='flex items-center space-x-2'>"
                        f"        {val_str} {status_badge}"
                        f"      </div>"
                        f"    </div>"
                    )
                
                header_badge = (
                    f"<span class='px-2 py-0.5 text-[9px] font-bold bg-neon-rose/10 text-neon-rose rounded border border-neon-rose/20 uppercase'>{overspent_count} Over Budget</span>"
                    if overspent_count > 0 else
                    f"<span class='px-2 py-0.5 text-[9px] font-bold bg-neon-green/10 text-neon-green rounded border border-neon-green/20 uppercase'>All Budgets Healthy</span>"
                )

                response = (
                    f"<div class='space-y-3'>"
                    f"  <div class='flex justify-between items-center'>"
                    f"    <strong class='text-xs text-slate-800 dark:text-white uppercase tracking-wider font-outfit'>Budget Health Audit</strong>"
                    f"    {header_badge}"
                    f"  </div>"
                    f"  <div class='space-y-2.5 max-h-52 overflow-y-auto pr-1'>"
                    f"    " + "\n".join(lines) + ""
                    f"  </div>"
                    f"</div>"
                )
            else:
                response = (
                    f"<div class='space-y-2'>"
                    f"  <strong class='text-xs text-slate-800 dark:text-white uppercase tracking-wider font-outfit'>No Active Limits</strong>"
                    f"  <p class='text-xs text-slate-600 dark:text-slate-300'>You have no category budgets set for this month. Set one by typing: <br><code class='block bg-slate-100 dark:bg-space-950 p-1.5 rounded mt-1 border border-slate-200 dark:border-space-800 text-[10px] text-neon-cyan font-mono'>set Food budget to 300</code></p>"
                    f"</div>"
                )
            return JsonResponse({'status': 'success', 'response': response})

        # Transaction Search query
        search_match = re.match(r'^(?:find|search|show|list)\s+(?:transactions|outflows|inflows|expenses|payments)?\s*(?:for|on|in|matching)?\s+(.+)$', q, re.IGNORECASE)
        if not search_match and (q_lower.startswith("find ") or q_lower.startswith("search ")):
            search_match = re.match(r'^(?:find|search)\s+(.+)$', q, re.IGNORECASE)
            
        if search_match:
            search_term = search_match.group(1).strip()
            txs = Transaction.objects.filter(user=user).filter(
                Q(category__icontains=search_term) | Q(description__icontains=search_term)
            )[:8]
            
            if txs.exists():
                lines = []
                for tx in txs:
                    type_indicator = "<span class='text-neon-green font-bold'>+</span>" if tx.transaction_type == 'IN' else "<span class='text-neon-rose font-bold'>-</span>"
                    amount_color = "text-neon-green" if tx.transaction_type == 'IN' else "text-slate-800 dark:text-white"
                    desc_str = f" <span class='text-slate-400 dark:text-slate-500'>({tx.description})</span>" if tx.description else ""
                    lines.append(
                        f"    <div class='flex justify-between items-center text-xs border-b border-slate-100 dark:border-space-850 pb-1.5'>"
                        f"      <div>"
                        f"        <strong class='text-slate-700 dark:text-slate-200'>{tx.category}</strong>{desc_str}"
                        f"        <p class='text-[10px] text-slate-400 dark:text-slate-500'>{tx.date.strftime('%b %d, %Y')}</p>"
                        f"      </div>"
                        f"      <div class='text-xs font-semibold {amount_color}'>"
                        f"        {type_indicator}{currency}{tx.amount:.2f}"
                        f"      </div>"
                        f"    </div>"
                    )
                response = (
                    f"<div class='space-y-3'>"
                    f"  <strong class='text-xs text-slate-800 dark:text-white uppercase tracking-wider font-outfit'>Ledger Query: '{search_term}'</strong>"
                    f"  <div class='space-y-2 max-h-52 overflow-y-auto pr-1'>"
                    f"    " + "\n".join(lines) + ""
                    f"  </div>"
                    f"</div>"
                )
            else:
                response = (
                    f"<div class='space-y-2'>"
                    f"  <strong class='text-xs text-slate-800 dark:text-white font-outfit uppercase tracking-wider'>No Matches Found</strong>"
                    f"  <p class='text-xs text-slate-600 dark:text-slate-300'>No ledger records match your query term <strong>'{search_term}'</strong> in category names or descriptions.</p>"
                    f"</div>"
                )
            return JsonResponse({'status': 'success', 'response': response})

        # Forecast Query
        if re.search(r'(?:predict|projection|forecast|future|extrapolate)', q_lower):
            savings_rate = int((balance / total_income) * 100) if total_income > 0 and balance > 0 else 0
            monthly_savings = balance if balance > 0 else 0.0
            
            p3 = monthly_savings * 3
            p6 = monthly_savings * 6
            p12 = monthly_savings * 12
            
            response = (
                f"<div class='space-y-2.5'>"
                f"  <strong class='text-xs text-slate-800 dark:text-white uppercase tracking-wider font-outfit'>Savings Projections</strong>"
                f"  <p class='text-[11px] text-slate-600 dark:text-slate-300'>Extrapolating your current monthly net surplus of <strong>{currency}{monthly_savings:.2f}</strong> ({savings_rate}% savings rate):</p>"
                f"  <ul class='space-y-1.5 pl-1'>"
                f"    <li class='flex justify-between items-center text-xs'>"
                f"      <span class='text-slate-500 dark:text-slate-400'>3-Month Cumulative:</span>"
                f"      <strong class='text-neon-cyan font-semibold'>{currency}{p3:.2f}</strong>"
                f"    </li>"
                f"    <li class='flex justify-between items-center text-xs'>"
                f"      <span class='text-slate-500 dark:text-slate-400'>6-Month Cumulative:</span>"
                f"      <strong class='text-neon-purple font-semibold'>{currency}{p6:.2f}</strong>"
                f"    </li>"
                f"    <li class='flex justify-between items-center text-xs'>"
                f"      <span class='text-slate-500 dark:text-slate-400'>12-Month Cumulative:</span>"
                f"      <strong class='text-neon-green font-semibold'>{currency}{p12:.2f}</strong>"
                f"    </li>"
                f"  </ul>"
                f"  <p class='text-[10px] text-slate-400 dark:text-slate-500 italic mt-2'>Note: Projections assume cashflow remains uniform. Configure category budgets to prevent leakage.</p>"
                f"</div>"
            )
            return JsonResponse({'status': 'success', 'response': response})

        # Fallback & General Cashflow Analysis (Detailed 50/30/20 breakdown)
        savings_rate = int((balance / total_income) * 100) if total_income > 0 and balance > 0 else 0
        
        tip_text = ""
        highest_cat = max(category_totals, key=category_totals.get) if category_totals else None
        highest_val = category_totals.get(highest_cat, 0.0) if highest_cat else 0.0
        if highest_cat and highest_val > 0:
            pct = int((highest_val / total_expense) * 100) if total_expense > 0 else 0
            tip_text = f"Your highest outflow category is <strong>{highest_cat}</strong> ({currency}{highest_val:.2f}, {pct}% of expenses). Placing a strict monthly limit here is your best saving leverage."
        else:
            tip_text = "Logging category expenses is the first step. Allocate a target budget to structure your outflows."

        needs_sum = 0.0
        wants_sum = 0.0
        for cat, val in category_totals.items():
            c_l = cat.lower()
            if any(n in c_l for n in ('rent', 'utility', 'utilities', 'food', 'bill', 'groceries')):
                needs_sum += val
            else:
                wants_sum += val
                
        needs_pct = int((needs_sum / total_income) * 100) if total_income > 0 else 0
        wants_pct = int((wants_sum / total_income) * 100) if total_income > 0 else 0
        savings_pct = savings_rate
        
        fallback_note = ""
        if not re.search(r'(?:save|saving|how to|cashflow|summary|status|health|analysis)', q_lower):
            fallback_note = f"<div class='text-[10px] text-slate-400 dark:text-slate-500 border-b border-slate-100 dark:border-space-850 pb-1 mb-1 italic'>Analyzed monthly cashflow. (Type 'help' to see my full commands list!)</div>"

        response = (
            f"{fallback_note}"
            f"<div class='space-y-2 text-xs'>"
            f"  <div class='flex justify-between items-center'>"
            f"    <strong class='text-xs text-slate-800 dark:text-white uppercase tracking-wider font-outfit'>Cashflow Diagnostic</strong>"
            f"    <span class='px-2 py-0.5 text-[9px] font-bold bg-neon-cyan/10 text-neon-cyan rounded border border-neon-cyan/20 uppercase'>{savings_rate}% Savings Rate</span>"
            f"  </div>"
            f"  <p class='text-slate-600 dark:text-slate-300'>Current Cashflow: Inflow <strong>{currency}{total_income:.2f}</strong> | Outflow <strong>{currency}{total_expense:.2f}</strong> | Net <strong>{currency}{balance:.2f}</strong>.</p>"
            f"  <div class='space-y-1.5 pt-1'>"
            f"    <strong class='text-[10px] text-slate-500 uppercase tracking-wide'>50/30/20 Allocation:</strong>"
            f"    <div class='grid grid-cols-3 gap-2 text-center text-[10px] font-semibold'>"
            f"      <div class='p-1 bg-slate-100 dark:bg-space-900 border border-slate-200 dark:border-space-800 rounded text-slate-700 dark:text-slate-300'>Needs: {needs_pct}% <span class='text-slate-400 dark:text-slate-500 font-normal'>(tgt 50%)</span></div>"
            f"      <div class='p-1 bg-slate-100 dark:bg-space-900 border border-slate-200 dark:border-space-800 rounded text-slate-700 dark:text-slate-300'>Wants: {wants_pct}% <span class='text-slate-400 dark:text-slate-500 font-normal'>(tgt 30%)</span></div>"
            f"      <div class='p-1 bg-slate-100 dark:bg-space-900 border border-slate-200 dark:border-space-800 rounded text-slate-700 dark:text-slate-300'>Savings: {savings_pct}% <span class='text-slate-400 dark:text-slate-500 font-normal'>(tgt 20%)</span></div>"
            f"    </div>"
            f"  </div>"
            f"  <p class='text-slate-600 dark:text-slate-300 text-[11px] leading-normal pt-1'>{tip_text}</p>"
            f"</div>"
        )
        return JsonResponse({'status': 'success', 'response': response})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)

# ----------------- Settings Views -----------------

@login_required
def toggle_theme(request):
    if request.method == 'POST':
        profile = request.user.profile
        profile.theme = 'light' if profile.theme == 'dark' else 'dark'
        profile.save()
        return JsonResponse({'status': 'success', 'theme': profile.theme})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)


@login_required
def update_currency(request):
    if request.method == 'POST':
        currency = request.POST.get('currency')
        rate_str = request.POST.get('conversion_rate', '1.0')
        if currency in ['$', '€', '₹', '£']:
            try:
                rate_str = rate_str.replace(',', '.')
                rate = Decimal(rate_str)
            except Exception as e:
                print(f"[Currency Convert Error] Failed to parse rate '{rate_str}': {e}")
                rate = Decimal('1.00')

            profile = request.user.profile
            
            # If a conversion rate is supplied and it's not 1.0, scale all existing amounts
            if rate != Decimal('1.00') and rate > 0:
                # 1. Transactions
                txs = Transaction.objects.filter(user=request.user)
                for tx in txs:
                    tx.amount = tx.amount * rate
                    tx.save()
                    
                # 2. Budgets
                budgets = Budget.objects.filter(user=request.user)
                for b in budgets:
                    b.amount = b.amount * rate
                    b.save()
                    
                # 3. Reminders
                reminders = Reminder.objects.filter(user=request.user)
                for r in reminders:
                    r.amount = r.amount * rate
                    r.save()
                    
                messages.success(request, f"Currency updated to {dict(Profile.CURRENCY_CHOICES).get(currency)}. Telemetry scaled by {rate}!")
            else:
                messages.success(request, f"Currency updated to {dict(Profile.CURRENCY_CHOICES).get(currency)} (values preserved).")

            profile.currency = currency
            profile.save()
        else:
            messages.error(request, "Invalid currency choice")
    return redirect('dashboard')

# ----------------- Action Views (CRUD) -----------------

def check_and_send_budget_alerts(user, category, tx_date):
    if not user.email:
        return
        
    try:
        if isinstance(tx_date, str):
            from datetime import datetime
            d = datetime.strptime(tx_date, '%Y-%m-%d').date()
        else:
            d = tx_date
            
        start_of_month = date(d.year, d.month, 1)
        _, last_day = calendar.monthrange(d.year, d.month)
        end_of_month = date(d.year, d.month, last_day)
        
        budgets = Budget.objects.filter(user=user, year=d.year, month=d.month, category__in=[category, 'Total'])
        if not budgets.exists():
            return
            
        month_txs = Transaction.objects.filter(user=user, date__range=(start_of_month, end_of_month), transaction_type='OUT')
        cat_expenses = month_txs.filter(category=category).aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        total_expenses = month_txs.aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        
        currency = user.profile.currency
        
        for b in budgets:
            limit = b.amount
            spent = total_expenses if b.category == 'Total' else cat_expenses
                
            if spent > limit:
                subject = f"BUDGET BREACH ALERT: {b.category} Budget Exceeded!"
                
                content_html = f"""
                <div style="background-color: #ff3b30; color: #ffffff; padding: 10px; border-radius: 8px; margin-bottom: 16px; font-weight: 700; font-size: 11px; text-align: center; text-transform: uppercase; letter-spacing: 0.5px; font-family: -apple-system, sans-serif;">
                    Target Category Breach Alert
                </div>
                <div style="background-color: #ffffff; padding: 12px; border-radius: 8px; border: 1px solid #e5e5ea; margin-bottom: 12px; box-shadow: 0 2px 8px rgba(0,0,0,0.02);">
                    <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse: collapse;">
                        <tr style="border-bottom: 1px solid #e5e5ea;">
                            <td style="padding: 8px 0; font-size: 11px; color: #8e8e93; font-family: -apple-system, sans-serif;">Target Category</td>
                            <td style="padding: 8px 0; font-size: 11px; color: #1c1c1e; font-weight: 600; text-align: right; font-family: -apple-system, sans-serif;">{b.category}</td>
                        </tr>
                        <tr style="border-bottom: 1px solid #e5e5ea;">
                            <td style="padding: 8px 0; font-size: 11px; color: #8e8e93; font-family: -apple-system, sans-serif;">Monthly Limit</td>
                            <td style="padding: 8px 0; font-size: 11px; color: #1c1c1e; font-weight: 600; text-align: right; font-family: -apple-system, sans-serif;">{currency}{limit:.2f}</td>
                        </tr>
                        <tr style="border-bottom: 1px solid #e5e5ea;">
                            <td style="padding: 8px 0; font-size: 11px; color: #8e8e93; font-family: -apple-system, sans-serif;">Accumulated Total</td>
                            <td style="padding: 8px 0; font-size: 11px; color: #ff3b30; font-weight: 600; text-align: right; font-family: -apple-system, sans-serif;">{currency}{spent:.2f}</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px 0; font-size: 11px; color: #8e8e93; font-family: -apple-system, sans-serif;">Threshold Overage</td>
                            <td style="padding: 8px 0; font-size: 11px; color: #ff3b30; font-weight: 600; text-align: right; font-family: -apple-system, sans-serif;">{currency}{spent - limit:.2f}</td>
                        </tr>
                    </table>
                </div>
                <p style="font-size: 11px; color: #8e8e93; line-height: 1.5; margin-top: 12px; margin-bottom: 0; font-family: -apple-system, sans-serif;">
                    Your account has exceeded the designated category thresholds for the month of {d.strftime('%B %Y')}. We recommend reviewing your transactions to reduce your outflows.
                </p>
                """
                
                email_body = render_html_email(
                    title=f"Category '{b.category}' Exceeded Limit",
                    subtitle="Budget Breach Transmission",
                    content_html=content_html
                )
                try:
                    send_user_mail(user, subject, email_body, is_html=True)
                except Exception as mail_err:
                    print(f"Error sending budget alert mail: {mail_err}")
    except Exception as e:
        print(f"Error in check_and_send_budget_alerts: {e}")


def add_transaction(request):
    if request.method == 'POST':
        amount = request.POST.get('amount')
        tx_type = request.POST.get('transaction_type')
        category = request.POST.get('category')
        tx_date = request.POST.get('date') or date.today().strftime('%Y-%m-%d')
        description = request.POST.get('description', '')
        
        try:
            tx = Transaction.objects.create(
                user=request.user,
                amount=Decimal(amount),
                transaction_type=tx_type,
                category=category,
                date=tx_date,
                description=description
            )
            messages.success(request, "Transaction successfully added!")
            
            # Send budget alert emails if transaction is an Expense (OUT)
            if tx_type == 'OUT':
                check_and_send_budget_alerts(request.user, category, tx_date)
        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f"Error saving transaction: {e}")
            
    return redirect('dashboard')


@login_required
def delete_transaction(request, pk):
    tx = get_object_or_404(Transaction, pk=pk, user=request.user)
    tx.delete()
    messages.success(request, "Transaction successfully deleted!")
    return redirect('dashboard')


@login_required
def set_budget(request):
    if request.method == 'POST':
        category = request.POST.get('category')
        amount = request.POST.get('amount')
        
        today = date.today()
        try:
            budget, created = Budget.objects.update_or_create(
                user=request.user,
                category=category,
                period='MONTHLY',
                month=today.month,
                year=today.year,
                defaults={'amount': Decimal(amount)}
            )
            messages.success(request, f"Budget set for {category}!")
        except Exception as e:
            messages.error(request, f"Error setting budget: {e}")
            
    return redirect('dashboard')


@login_required
def add_reminder(request):
    if request.method == 'POST':
        title = request.POST.get('title')
        amount = request.POST.get('amount')
        due_date = request.POST.get('due_date')
        is_recurring = request.POST.get('is_recurring') == 'on'
        recurrence_period = request.POST.get('recurrence_period') if is_recurring else None
        
        try:
            Reminder.objects.create(
                user=request.user,
                title=title,
                amount=Decimal(amount),
                due_date=due_date,
                is_recurring=is_recurring,
                recurrence_period=recurrence_period
            )
            messages.success(request, "Bill reminder successfully added!")
        except Exception as e:
            messages.error(request, f"Error adding reminder: {e}")
            
    return redirect('dashboard')


@login_required
def pay_reminder(request, pk):
    reminder = get_object_or_404(Reminder, pk=pk, user=request.user)
    
    if reminder.is_recurring and reminder.recurrence_period:
        # Create a new reminder for the next period
        current_due = reminder.due_date
        next_due = current_due
        
        if reminder.recurrence_period == 'DAILY':
            next_due = current_due + timedelta(days=1)
        elif reminder.recurrence_period == 'WEEKLY':
            next_due = current_due + timedelta(weeks=1)
        elif reminder.recurrence_period == 'MONTHLY':
            # Add one month handling days overflow
            month = current_due.month
            year = current_due.year
            if month == 12:
                month = 1
                year += 1
            else:
                month += 1
            _, last_day = calendar.monthrange(year, month)
            day = min(current_due.day, last_day)
            next_due = date(year, month, day)
        elif reminder.recurrence_period == 'YEARLY':
            # Add one year
            year = current_due.year + 1
            month = current_due.month
            _, last_day = calendar.monthrange(year, month)
            day = min(current_due.day, last_day)
            next_due = date(year, month, day)
            
        # Spawn next recurrence
        Reminder.objects.create(
            user=reminder.user,
            title=reminder.title,
            amount=reminder.amount,
            due_date=next_due,
            is_recurring=True,
            recurrence_period=reminder.recurrence_period,
            is_paid=False
        )
        
    reminder.is_paid = True
    reminder.save()
    messages.success(request, f"Marked '{reminder.title}' as Paid!")
    return redirect('dashboard')

# ----------------- Helper: Seed Demo Data -----------------

def seed_user_demo_data(user):
    """
    Seeds rich, colorful, realistic transaction, budget and reminder data
    so the user sees a beautiful, fully functional dashboard right away.
    """
    today = date.today()
    # Income Transactions
    Transaction.objects.create(user=user, amount=Decimal('4200.00'), transaction_type='IN', category='Salary', date=today - timedelta(days=5), description='Monthly Tech Lead Salary')
    Transaction.objects.create(user=user, amount=Decimal('850.00'), transaction_type='IN', category='Freelance', date=today - timedelta(days=12), description='Web Design Project Delivery')
    Transaction.objects.create(user=user, amount=Decimal('120.00'), transaction_type='IN', category='Investment', date=today - timedelta(days=15), description='Stock Dividends')

    # Expenses Current Month
    Transaction.objects.create(user=user, amount=Decimal('1500.00'), transaction_type='OUT', category='Rent', date=today - timedelta(days=8), description='Studio Apartment Rent')
    Transaction.objects.create(user=user, amount=Decimal('245.50'), transaction_type='OUT', category='Food', date=today - timedelta(days=1), description='Dinner & Drinks with Friends')
    Transaction.objects.create(user=user, amount=Decimal('85.00'), transaction_type='OUT', category='Food', date=today - timedelta(days=4), description='Weekly Groceries')
    Transaction.objects.create(user=user, amount=Decimal('115.00'), transaction_type='OUT', category='Utilities', date=today - timedelta(days=7), description='High-speed Fiber Internet & Power')
    Transaction.objects.create(user=user, amount=Decimal('120.00'), transaction_type='OUT', category='Entertainment', date=today - timedelta(days=3), description='Concert Tickets')
    Transaction.objects.create(user=user, amount=Decimal('65.00'), transaction_type='OUT', category='Travel', date=today - timedelta(days=2), description='Gas & Commute')
    Transaction.objects.create(user=user, amount=Decimal('45.00'), transaction_type='OUT', category='Other', date=today - timedelta(days=10), description='Tech Magazine Subscription')

    # Seed Last Month Transactions to make recommendations look incredibly smart
    last_month_first = date(today.year, today.month, 1) - timedelta(days=15)
    Transaction.objects.create(user=user, amount=Decimal('4200.00'), transaction_type='IN', category='Salary', date=last_month_first, description='Last Month Salary')
    Transaction.objects.create(user=user, amount=Decimal('1500.00'), transaction_type='OUT', category='Rent', date=last_month_first + timedelta(days=1), description='Studio Apartment Rent')
    Transaction.objects.create(user=user, amount=Decimal('180.00'), transaction_type='OUT', category='Food', date=last_month_first + timedelta(days=5), description='Last Month Food Total (lower)')
    Transaction.objects.create(user=user, amount=Decimal('60.00'), transaction_type='OUT', category='Entertainment', date=last_month_first + timedelta(days=10), description='Movie tickets')

    # Budgets for Current Month
    Budget.objects.create(user=user, category='Food', amount=Decimal('350.00'), period='MONTHLY', month=today.month, year=today.year)
    Budget.objects.create(user=user, category='Rent', amount=Decimal('1500.00'), period='MONTHLY', month=today.month, year=today.year)
    Budget.objects.create(user=user, category='Utilities', amount=Decimal('150.00'), period='MONTHLY', month=today.month, year=today.year)
    Budget.objects.create(user=user, category='Entertainment', amount=Decimal('100.00'), period='MONTHLY', month=today.month, year=today.year) # Will trigger warning (spent 120/100, danger)
    Budget.objects.create(user=user, category='Total', amount=Decimal('2800.00'), period='MONTHLY', month=today.month, year=today.year)

    # Reminders
    Reminder.objects.create(user=user, title='AWS Cloud Hosting', amount=Decimal('42.15'), due_date=today + timedelta(days=5), is_recurring=True, recurrence_period='MONTHLY')
    Reminder.objects.create(user=user, title='Gym Membership', amount=Decimal('60.00'), due_date=today + timedelta(days=11), is_recurring=True, recurrence_period='MONTHLY')
    Reminder.objects.create(user=user, title='Car Insurance', amount=Decimal('150.00'), due_date=today + timedelta(days=20), is_recurring=False)


# ----------------- Additional Action Views (Phase 2) -----------------

@login_required
def delete_budget_category(request, category):
    today = date.today()
    deleted_count, _ = Budget.objects.filter(
        user=request.user, 
        category=category, 
        year=today.year, 
        month=today.month
    ).delete()
    
    if deleted_count > 0:
        messages.success(request, f"Budget limit for {category} cleared successfully.")
    else:
        messages.info(request, f"No budget set for {category}.")
    return redirect('dashboard')


@login_required
def export_csv_view(request):
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="leoxur_financial_export.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['LEOXUR FINANCIAL TELEMETRY EXPORT'])
    writer.writerow(['User', request.user.username])
    writer.writerow(['Export Date', date.today().strftime('%Y-%m-%d')])
    writer.writerow([])
    
    writer.writerow(['--- TRANSACTIONS LEDGER ---'])
    writer.writerow(['ID', 'Description', 'Type', 'Category', 'Date', 'Amount'])
    txs = Transaction.objects.filter(user=request.user).order_by('-date')
    for tx in txs:
        writer.writerow([tx.id, tx.description, tx.get_transaction_type_display(), tx.category, tx.date.strftime('%Y-%m-%d'), tx.amount])
        
    writer.writerow([])
    writer.writerow(['--- MONTHLY BUDGET TARGETS ---'])
    writer.writerow(['Category', 'Allowance Amount', 'Month', 'Year'])
    budgets = Budget.objects.filter(user=request.user)
    for b in budgets:
        writer.writerow([b.category, b.amount, b.month, b.year])
        
    writer.writerow([])
    writer.writerow(['--- SCHEDULED BILL REMINDERS ---'])
    writer.writerow(['Title', 'Amount', 'Due Date', 'Is Recurring', 'Frequency', 'Status'])
    reminders = Reminder.objects.filter(user=request.user)
    for r in reminders:
        status = 'Paid' if r.is_paid else 'Pending'
        writer.writerow([r.title, r.amount, r.due_date.strftime('%Y-%m-%d'), r.is_recurring, r.get_recurrence_period_display() if r.is_recurring else 'N/A', status])
        
    return response


def generate_excel_data(user):
    from openpyxl import Workbook
    import io
    wb = Workbook()
    
    ws1 = wb.active
    ws1.title = "Transactions Ledger"
    ws1.append(['Transaction ID', 'Description', 'Type', 'Category', 'Date', 'Amount'])
    txs = Transaction.objects.filter(user=user).order_by('-date')
    for tx in txs:
        ws1.append([tx.id, tx.description, tx.get_transaction_type_display(), tx.category, tx.date.strftime('%Y-%m-%d'), float(tx.amount)])
        
    ws2 = wb.create_sheet(title="Budget Targets")
    ws2.append(['Category', 'Limit Amount', 'Month', 'Year'])
    budgets = Budget.objects.filter(user=user)
    for b in budgets:
        ws2.append([b.category, float(b.amount), b.month, b.year])
        
    ws3 = wb.create_sheet(title="Bill Reminders")
    ws3.append(['Title', 'Amount', 'Due Date', 'Is Recurring', 'Recurrence Frequency', 'Status'])
    reminders = Reminder.objects.filter(user=user)
    for r in reminders:
        status = 'Paid' if r.is_paid else 'Pending'
        ws3.append([r.title, float(r.amount), r.due_date.strftime('%Y-%m-%d'), "Yes" if r.is_recurring else "No", r.get_recurrence_period_display() if r.is_recurring else 'N/A', status])
        
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    data = excel_buffer.getvalue()
    excel_buffer.close()
    return data


def generate_pdf_data(user, total_income, total_expense, balance, currency):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    import io
    
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    story = []
    
    styles = getSampleStyleSheet()
    primary_color = colors.HexColor("#1d1d1f")
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=22,
        textColor=primary_color,
        spaceAfter=10
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.gray,
        spaceAfter=20
    )
    
    h2_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=primary_color,
        spaceBefore=12,
        spaceAfter=8
    )
    
    story.append(Paragraph("LEOXUR FINANCIAL LEDGER", title_style))
    story.append(Paragraph(f"Monthly Telemetry Statement for node '{user.username}' | Date: {date.today().strftime('%B %d, %Y')}", subtitle_style))
    story.append(Spacer(1, 5))
    
    overview_data = [
        ["Telemetry Metric", "Value"],
        ["Total Earnings (Inflow)", f"{currency}{total_income:.2f}"],
        ["Total Spending (Outflow)", f"{currency}{total_expense:.2f}"],
        ["Current Monthly Balance", f"{currency}{balance:.2f}"]
    ]
    overview_table = Table(overview_data, colWidths=[200, 150])
    overview_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), primary_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('BOTTOMPADDING', (0,0), (-1,0), 4),
        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor("#f8fafc")),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 9),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    
    story.append(Paragraph("Monthly Overview Summary", h2_style))
    story.append(overview_table)
    story.append(Spacer(1, 15))
    
    story.append(Paragraph("Recent Transactions Statement", h2_style))
    tx_headers = ["ID", "Description", "Type", "Category", "Date", "Amount"]
    tx_rows = [tx_headers]
    
    txs = Transaction.objects.filter(user=user).order_by('-date')[:30]
    for tx in txs:
        tx_rows.append([
            f"#{tx.id:05d}",
            tx.description or "N/A",
            tx.get_transaction_type_display(),
            tx.category,
            tx.date.strftime('%Y-%m-%d'),
            f"{'+' if tx.transaction_type == 'IN' else '-'}{currency}{tx.amount:.2f}"
        ])
        
    tx_table = Table(tx_rows, colWidths=[50, 150, 60, 90, 80, 70])
    tx_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), primary_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 8),
        ('BOTTOMPADDING', (0,0), (-1,0), 4),
        ('ALIGN', (-1,0), (-1,-1), 'RIGHT'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#e2e8f0")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('PADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(tx_table)
    
    doc.build(story)
    data = pdf_buffer.getvalue()
    pdf_buffer.close()
    return data


def render_html_email(title, subtitle, content_html):
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{
                margin: 0;
                padding: 12px;
                background-color: #f5f5f7;
                color: #1c1c1e;
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            }}
        </style>
    </head>
    <body style="margin: 0; padding: 12px; background-color: #f5f5f7; color: #1c1c1e;">
        <table width="100%" border="0" cellspacing="0" cellpadding="0" style="background-color: #f5f5f7; padding: 20px 0;">
            <tr>
                <td align="center" valign="top">
                    <table width="100%" max-width="600" style="width: 100%; max-width: 600px; background-color: #ffffff; border: 1px solid #e5e5ea; border-radius: 16px; border-collapse: separate; overflow: hidden; box-shadow: 0 4px 24px rgba(0,0,0,0.06);" cellspacing="0" cellpadding="0">
                        <!-- Header -->
                        <tr>
                            <td style="padding: 20px 24px; border-bottom: 1px solid #e5e5ea; background-color: #fafafa;">
                                <table width="100%" cellspacing="0" cellpadding="0">
                                    <tr>
                                        <td>
                                            <h2 style="margin: 0; font-size: 11px; font-weight: 700; color: #a855f7; letter-spacing: 1px; font-family: -apple-system, sans-serif; text-transform: uppercase;">LEOXUR TELEMETRY</h2>
                                            <div style="font-size: 8px; color: #8e8e93; margin-top: 2px; text-transform: uppercase; letter-spacing: 0.5px;">{subtitle}</div>
                                        </td>
                                    </tr>
                                </table>
                            </td>
                        </tr>
                        <!-- Content -->
                        <tr>
                            <td style="padding: 24px 30px;">
                                <h1 style="margin: 0 0 16px 0; font-size: 18px; font-weight: 600; color: #1c1c1e; letter-spacing: -0.3px;">{title}</h1>
                                {content_html}
                            </td>
                        </tr>
                        <!-- Footer -->
                        <tr>
                            <td style="padding: 14px 24px; background-color: #fafafa; border-top: 1px solid #e5e5ea; text-align: center; font-size: 8px; color: #8e8e93;">
                                Automated transmission. Settings and UI are synchronized.
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """


def render_compact_transaction_email(category_name, amount, tx_type_str, date_str, description, currency, tx_type):
    flow_color = '#34c759' if tx_type == 'IN' else '#ff3b30'
    flow_label = 'INCOME' if tx_type == 'IN' else 'EXPENSE'
    
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <style>
            body {{
                margin: 0;
                padding: 0;
                background-color: #f5f5f7;
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            }}
        </style>
    </head>
    <body style="margin: 0; padding: 10px; background-color: #f5f5f7;">
        <table width="100%" border="0" cellspacing="0" cellpadding="0" style="background-color: #f5f5f7;">
            <tr>
                <td align="center" valign="top">
                    <table width="100%" max-width="440" style="width: 100%; max-width: 440px; background-color: #ffffff; border: 1px solid #e5e5ea; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 12px rgba(0,0,0,0.05); border-collapse: separate;" cellspacing="0" cellpadding="0">
                        <!-- Banner Header -->
                        <tr>
                            <td style="background-color: {flow_color}; color: #ffffff; padding: 8px 12px; font-size: 10px; font-weight: 700; letter-spacing: 0.5px; text-transform: uppercase; text-align: center;">
                                Transaction Processed: {flow_label}
                            </td>
                        </tr>
                        <!-- Details Table -->
                        <tr>
                            <td style="padding: 16px;">
                                <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse: collapse;">
                                    <tr>
                                        <td style="font-size: 8px; text-transform: uppercase; color: #8e8e93; padding-bottom: 2px;">Category</td>
                                        <td style="font-size: 8px; text-transform: uppercase; color: #8e8e93; padding-bottom: 2px; text-align: center;">Amount</td>
                                        <td style="font-size: 8px; text-transform: uppercase; color: #8e8e93; padding-bottom: 2px; text-align: right;">Date</td>
                                    </tr>
                                    <tr>
                                        <td style="font-size: 12px; color: #1c1c1e; font-weight: 600; padding-right: 8px;">{category_name}</td>
                                        <td style="font-size: 12px; color: #1c1c1e; font-weight: 600; text-align: center; padding-right: 8px;">{currency}{amount:.2f}</td>
                                        <td style="font-size: 12px; color: #1c1c1e; font-weight: 600; text-align: right;">{date_str}</td>
                                    </tr>
                                </table>
                                <div style="margin-top: 10px; border-top: 1px solid #e5e5ea; padding-top: 8px; font-size: 10px; color: #8e8e93;">
                                    <strong>Note:</strong> {description or 'N/A'}
                                </div>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """


def send_user_mail(user, subject, body, attachments=None, is_html=False):
    from django.core.mail import get_connection, EmailMultiAlternatives
    import re
    profile = user.profile
    if profile.smtp_host and profile.smtp_user and profile.smtp_password:
        connection = get_connection(
            backend='django.core.mail.backends.smtp.EmailBackend',
            host=profile.smtp_host,
            port=profile.smtp_port,
            username=profile.smtp_user,
            password=profile.smtp_password,
            use_tls=True,
        )
        from_email = f"Leoxur Financial <{profile.smtp_user}>"
    else:
        connection = None
        from_email = None

    if is_html:
        plain_text = re.sub(r'<[^>]+>', '', body)
    else:
        plain_text = body

    email = EmailMultiAlternatives(
        subject=subject,
        body=plain_text,
        from_email=from_email,
        to=[user.email],
        connection=connection
    )
    email.extra_headers['X-Leoxur-Auto'] = 'True'
    email.extra_headers['Auto-Submitted'] = 'auto-generated'
    if is_html:
        email.attach_alternative(body, "text/html")
        
    if attachments:
        for name, content, mimetype in attachments:
            email.attach(name, content, mimetype)
    email.send()


@login_required
def update_email_settings(request):
    if request.method == 'POST':
        profile = request.user.profile
        profile.smtp_host = request.POST.get('smtp_host', 'smtp.gmail.com')
        profile.smtp_port = int(request.POST.get('smtp_port', '587'))
        profile.smtp_user = request.POST.get('smtp_user', '')
        profile.smtp_password = request.POST.get('smtp_password', '')
        profile.imap_host = request.POST.get('imap_host', 'imap.gmail.com')
        profile.imap_port = int(request.POST.get('imap_port', '993'))
        profile.imap_user = request.POST.get('imap_user', '')
        profile.imap_password = request.POST.get('imap_password', '')
        profile.auto_fetch_emails = request.POST.get('auto_fetch_emails') == 'on'
        profile.save()
        
        # Update User email profile
        user = request.user
        user.email = request.POST.get('account_email', '')
        user.save()
        
        messages.success(request, "Email SMTP/IMAP settings updated successfully!")
    return redirect('dashboard')


@login_required
def test_email_report(request):
    user = request.user
    if not user.email:
        messages.error(request, "Please register an email address on your user profile before running tests.")
        return redirect('dashboard')
        
    today = date.today()
    start_of_month = date(today.year, today.month, 1)
    _, last_day = calendar.monthrange(today.year, today.month)
    end_of_month = date(today.year, today.month, last_day)

    month_txs = Transaction.objects.filter(user=user, date__range=(start_of_month, end_of_month))
    total_income = month_txs.filter(transaction_type='IN').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    total_expense = month_txs.filter(transaction_type='OUT').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    balance = total_income - total_expense

    raw_currency = user.profile.currency
    currency_map = {
        '$': '$',
        '€': 'EUR ',
        '₹': 'INR ',
        '£': 'GBP ',
    }
    currency = currency_map.get(raw_currency, raw_currency + ' ')

    budgets = Budget.objects.filter(user=user, year=today.year, month=today.month)
    
    expense_pct = int((total_expense / total_income * 100)) if total_income > 0 else 0
    if expense_pct > 100: expense_pct = 100

    balances_html = f"""
    <table width="100%" cellspacing="0" cellpadding="0" style="margin-bottom: 24px; border-spacing: 12px; border-collapse: separate; margin-left: -12px; margin-right: -12px;">
        <tr>
            <td width="50%" style="background-color: #ffffff; padding: 18px; border-radius: 12px; border: 1px solid #e5e5ea; font-family: -apple-system, sans-serif; box-shadow: 0 4px 10px rgba(0,0,0,0.02);">
                <div style="font-size: 9px; text-transform: uppercase; color: #8e8e93; letter-spacing: 0.5px; font-weight: 700; margin-bottom: 4px;">Total Inflows</div>
                <div style="font-size: 20px; font-weight: 700; color: #34c759; margin-top: 6px;">{currency}{total_income:.2f}</div>
            </td>
            <td width="50%" style="background-color: #ffffff; padding: 18px; border-radius: 12px; border: 1px solid #e5e5ea; font-family: -apple-system, sans-serif; box-shadow: 0 4px 10px rgba(0,0,0,0.02);">
                <div style="font-size: 9px; text-transform: uppercase; color: #8e8e93; letter-spacing: 0.5px; font-weight: 700; margin-bottom: 4px;">Total Outflows</div>
                <div style="font-size: 20px; font-weight: 700; color: #ff3b30; margin-top: 6px;">{currency}{total_expense:.2f}</div>
            </td>
        </tr>
    </table>
    
    <div style="background-color: #ffffff; padding: 18px; border-radius: 12px; border: 1px solid #e5e5ea; margin-bottom: 24px; font-family: -apple-system, sans-serif; box-shadow: 0 4px 10px rgba(0,0,0,0.02);">
        <table width="100%" cellspacing="0" cellpadding="0" style="margin-bottom: 10px; border-collapse: collapse;">
            <tr>
                <td style="font-size: 10px; text-transform: uppercase; color: #8e8e93; letter-spacing: 0.5px; font-weight: 700;">Net Monthly Balance</td>
                <td style="font-size: 16px; font-weight: 700; color: {'#a855f7' if balance >= 0 else '#ff3b30'}; text-align: right;">{currency}{balance:.2f}</td>
            </tr>
        </table>
        <div style="background-color: #e5e5ea; border-radius: 9999px; height: 6px; overflow: hidden; width: 100%; margin-top: 6px;">
            <div style="background-color: #a855f7; height: 100%; width: {expense_pct}%; border-radius: 9999px;"></div>
        </div>
        <div style="font-size: 8px; color: #8e8e93; margin-top: 6px; text-align: right;">Spending ratio: {expense_pct}%</div>
    </div>
    """

    budget_lines = []
    for b in budgets:
        spent = total_expense if b.category == 'Total' else month_txs.filter(category=b.category, transaction_type='OUT').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        pct = int((spent / b.amount * 100)) if b.amount > 0 else 0
        fill_pct = pct if pct <= 100 else 100
        bar_color = "#ff3b30" if pct >= 100 else "#a855f7"
        budget_lines.append(f"""
        <div style="margin-bottom: 16px; font-family: -apple-system, sans-serif;">
            <table width="100%" cellspacing="0" cellpadding="0" style="margin-bottom: 4px; border-collapse: collapse;">
                <tr>
                    <td style="font-size: 11px; font-weight: 600; color: #1c1c1e;">{b.category}</td>
                    <td style="font-size: 10px; color: #8e8e93; text-align: right;">{currency}{spent:.2f} / {currency}{b.amount:.2f} ({pct}%)</td>
                </tr>
            </table>
            <div style="background-color: #e5e5ea; border-radius: 9999px; height: 5px; overflow: hidden; width: 100%; border: 1px solid #e5e5ea;">
                <div style="background-color: {bar_color}; height: 100%; width: {fill_pct}%; border-radius: 9999px;"></div>
            </div>
        </div>
        """)
        
    budgets_html = f"""
    <div style="margin-bottom: 24px; font-family: -apple-system, sans-serif;">
        <h3 style="font-size: 11px; text-transform: uppercase; color: #a855f7; letter-spacing: 0.5px; margin-bottom: 12px; border-bottom: 1px solid #e5e5ea; padding-bottom: 6px; font-weight: 700;">Budget Limits</h3>
        {"".join(budget_lines) if budget_lines else '<div style="font-size: 9px; color: #8e8e93;">No active budget targets.</div>'}
    </div>
    """

    recent_rows = []
    txs = Transaction.objects.filter(user=user).order_by('-date')[:5]
    for tx in txs:
        amt_color = "#34c759" if tx.transaction_type == 'IN' else "#ff3b30"
        amt_sign = "+" if tx.transaction_type == 'IN' else "-"
        recent_rows.append(f"""
        <tr style="border-bottom: 1px solid #e5e5ea;">
            <td style="padding: 10px 0; font-size: 11px; color: #1c1c1e; font-family: -apple-system, sans-serif;">{tx.description or 'N/A'}</td>
            <td style="padding: 10px 0; font-size: 11px; color: #8e8e93; font-family: -apple-system, sans-serif;">{tx.category}</td>
            <td style="padding: 10px 0; font-size: 11px; text-align: right; font-weight: 600; color: {amt_color}; font-family: -apple-system, sans-serif;">{amt_sign}{currency}{tx.amount:.2f}</td>
        </tr>
        """)
        
    recent_html = f"""
    <div style="margin-bottom: 8px; font-family: -apple-system, sans-serif;">
        <h3 style="font-size: 11px; text-transform: uppercase; color: #a855f7; letter-spacing: 0.5px; margin-bottom: 12px; border-bottom: 1px solid #e5e5ea; padding-bottom: 6px; font-weight: 700;">Recent Statements</h3>
        <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse: collapse;">
            <thead>
                <tr style="text-align: left; border-bottom: 1px solid #e5e5ea;">
                    <th style="padding-bottom: 6px; font-size: 10px; text-transform: uppercase; color: #8e8e93; font-weight: 700; font-family: -apple-system, sans-serif;">Description</th>
                    <th style="padding-bottom: 6px; font-size: 10px; text-transform: uppercase; color: #8e8e93; font-weight: 700; font-family: -apple-system, sans-serif;">Category</th>
                    <th style="padding-bottom: 6px; font-size: 10px; text-transform: uppercase; color: #8e8e93; text-align: right; font-weight: 700; font-family: -apple-system, sans-serif;">Amount</th>
                </tr>
            </thead>
            <tbody>
                {"".join(recent_rows) if recent_rows else '<tr><td colspan="3" style="padding: 12px 0; font-size: 11px; color: #8e8e93; text-align: center; font-family: -apple-system, sans-serif;">No transactions found.</td></tr>'}
            </tbody>
        </table>
    </div>
    """

    content_html = f"""
    {balances_html}
    {budgets_html}
    {recent_html}
    """

    email_body = render_html_email(
        title="Daily Financial Summary Report",
        subtitle=f"Generated Statement for '{user.username}'",
        content_html=content_html
    )

    try:
        excel_data = generate_excel_data(user)
        pdf_data = generate_pdf_data(user, total_income, total_expense, balance, currency)
        
        send_user_mail(
            user,
            f"Daily Financial Report - {today.strftime('%b %d, %Y')}",
            email_body,
            attachments=[
                (f"Statement_{today.strftime('%Y%m%d')}.pdf", pdf_data, "application/pdf"),
                (f"Statement_{today.strftime('%Y%m%d')}.xlsx", excel_data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            ],
            is_html=True
        )
        messages.success(request, f"Test Daily Report email successfully sent to {user.email}!")
    except Exception as e:
        messages.error(request, f"Error sending test report email: {e}")
    return redirect('dashboard')


@login_required
def test_email_alert(request):
    user = request.user
    if not user.email:
        messages.error(request, "Please register an email address on your user profile before running tests.")
        return redirect('dashboard')
        
    currency = user.profile.currency
    
    subject = "BUDGET BREACH ALERT: Entertainment Budget Exceeded"
    
    content_html = f"""
    <div style="background-color: #ff3b30; color: #ffffff; padding: 10px; border-radius: 8px; margin-bottom: 16px; font-weight: 700; font-size: 11px; text-align: center; text-transform: uppercase; letter-spacing: 0.5px; font-family: -apple-system, sans-serif;">
        Target Category Breach Alert
    </div>
    <div style="background-color: #ffffff; padding: 12px; border-radius: 8px; border: 1px solid #e5e5ea; margin-bottom: 12px; box-shadow: 0 2px 8px rgba(0,0,0,0.02);">
        <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse: collapse;">
            <tr style="border-bottom: 1px solid #e5e5ea;">
                <td style="padding: 8px 0; font-size: 11px; color: #8e8e93; font-family: -apple-system, sans-serif;">Target Category</td>
                <td style="padding: 8px 0; font-size: 11px; color: #1c1c1e; font-weight: 600; text-align: right; font-family: -apple-system, sans-serif;">Entertainment</td>
            </tr>
            <tr style="border-bottom: 1px solid #e5e5ea;">
                <td style="padding: 8px 0; font-size: 11px; color: #8e8e93; font-family: -apple-system, sans-serif;">Monthly Limit</td>
                <td style="padding: 8px 0; font-size: 11px; color: #1c1c1e; font-weight: 600; text-align: right; font-family: -apple-system, sans-serif;">{currency}250.00</td>
            </tr>
            <tr style="border-bottom: 1px solid #e5e5ea;">
                <td style="padding: 8px 0; font-size: 11px; color: #8e8e93; font-family: -apple-system, sans-serif;">Accumulated Total</td>
                <td style="padding: 8px 0; font-size: 11px; color: #ff3b30; font-weight: 600; text-align: right; font-family: -apple-system, sans-serif;">{currency}320.00</td>
            </tr>
            <tr>
                <td style="padding: 8px 0; font-size: 11px; color: #8e8e93; font-family: -apple-system, sans-serif;">Threshold Overage</td>
                <td style="padding: 8px 0; font-size: 11px; color: #ff3b30; font-weight: 600; text-align: right; font-family: -apple-system, sans-serif;">{currency}70.00</td>
            </tr>
        </table>
    </div>
    <p style="font-size: 11px; color: #8e8e93; line-height: 1.5; margin-top: 12px; margin-bottom: 0; font-family: -apple-system, sans-serif;">
        Your account has exceeded the designated category thresholds for the month of {date.today().strftime('%B %Y')}. We recommend reviewing your transactions to reduce your outflows.
    </p>
    """
    
    email_body = render_html_email(
        title="Category 'Entertainment' Exceeded Limit",
        subtitle="Budget Breach Transmission",
        content_html=content_html
    )
    
    try:
        send_user_mail(user, subject, email_body, is_html=True)
        messages.success(request, f"Test Budget Alert email successfully sent to {user.email}!")
    except Exception as e:
        messages.error(request, f"Error sending test alert email: {e}")
    return redirect('dashboard')


@login_required
def test_email_reminder(request):
    user = request.user
    if not user.email:
        messages.error(request, "Please register an email address on your user profile before running tests.")
        return redirect('dashboard')
        
    currency = user.profile.currency
    
    subject = "BILL DUE TODAY: Internet & Broadband subscription"
    
    content_html = f"""
    <div style="background-color: #a855f7; color: #ffffff; padding: 10px; border-radius: 8px; margin-bottom: 16px; font-weight: 700; font-size: 11px; text-align: center; text-transform: uppercase; letter-spacing: 0.5px; font-family: -apple-system, sans-serif;">
        Scheduled Payment Due Today
    </div>
    <div style="background-color: #ffffff; padding: 12px; border-radius: 8px; border: 1px solid #e5e5ea; margin-bottom: 12px; box-shadow: 0 2px 8px rgba(0,0,0,0.02);">
        <table width="100%" cellspacing="0" cellpadding="0" style="border-collapse: collapse;">
            <tr style="border-bottom: 1px solid #e5e5ea;">
                <td style="padding: 8px 0; font-size: 11px; color: #8e8e93; font-family: -apple-system, sans-serif;">Bill Description</td>
                <td style="padding: 8px 0; font-size: 11px; color: #1c1c1e; font-weight: 600; text-align: right; font-family: -apple-system, sans-serif;">Internet & Broadband subscription</td>
            </tr>
            <tr style="border-bottom: 1px solid #e5e5ea;">
                <td style="padding: 8px 0; font-size: 11px; color: #8e8e93; font-family: -apple-system, sans-serif;">Amount Due</td>
                <td style="padding: 8px 0; font-size: 11px; color: #1c1c1e; font-weight: 600; text-align: right; font-family: -apple-system, sans-serif;">{currency}79.99</td>
            </tr>
            <tr style="border-bottom: 1px solid #e5e5ea;">
                <td style="padding: 8px 0; font-size: 11px; color: #8e8e93; font-family: -apple-system, sans-serif;">Due Date</td>
                <td style="padding: 8px 0; font-size: 11px; color: #1c1c1e; font-weight: 600; text-align: right; font-family: -apple-system, sans-serif;">{date.today().strftime('%A, %b %d, %Y')}</td>
            </tr>
            <tr>
                <td style="padding: 8px 0; font-size: 11px; color: #8e8e93; font-family: -apple-system, sans-serif;">Recurrence</td>
                <td style="padding: 8px 0; font-size: 11px; color: #1c1c1e; font-weight: 600; text-align: right; font-family: -apple-system, sans-serif;">Yes (Monthly)</td>
            </tr>
        </table>
    </div>
    <p style="font-size: 11px; color: #8e8e93; line-height: 1.5; margin-top: 12px; margin-bottom: 0; font-family: -apple-system, sans-serif;">
        This automated transmission has been dispatched to remind you that your scheduled payment is due today. Please log in to your dashboard to pay the reminder and update your records.
    </p>
    """
    
    email_body = render_html_email(
        title="Internet & Broadband Subscription Payment Due",
        subtitle="Scheduled Reminder Transmission",
        content_html=content_html
    )
    
    try:
        send_user_mail(user, subject, email_body, is_html=True)
        messages.success(request, f"Test Bill Reminder email successfully sent to {user.email}!")
    except Exception as e:
        messages.error(request, f"Error sending test bill reminder email: {e}")
    return redirect('dashboard')


@login_required
def export_excel_view(request):
    from django.http import HttpResponse
    try:
        data = generate_excel_data(request.user)
        response = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="leoxur_financial_export.xlsx"'
        return response
    except Exception as e:
        messages.error(request, f"Error generating Excel statement: {e}")
        return redirect('dashboard')


@login_required
def export_pdf_view(request):
    from django.http import HttpResponse
    try:
        today = date.today()
        start_of_month = date(today.year, today.month, 1)
        _, last_day = calendar.monthrange(today.year, today.month)
        end_of_month = date(today.year, today.month, last_day)
        
        month_txs = Transaction.objects.filter(user=request.user, date__range=(start_of_month, end_of_month))
        total_income = month_txs.filter(transaction_type='IN').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        total_expense = month_txs.filter(transaction_type='OUT').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
        balance = total_income - total_expense
        raw_currency = request.user.profile.currency
        currency_map = {
            '$': '$',
            '€': 'EUR ',
            '₹': 'INR ',
            '£': 'GBP ',
        }
        currency = currency_map.get(raw_currency, raw_currency + ' ')
        
        data = generate_pdf_data(request.user, total_income, total_expense, balance, currency)
        response = HttpResponse(data, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="leoxur_financial_statement.pdf"'
        return response
    except Exception as e:
        messages.error(request, f"Error generating PDF statement: {e}")
        return redirect('dashboard')


# ----------------- Custom Category & modifiers (Phase 3) -----------------

@login_required
def add_category(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        is_income = request.POST.get('is_income') == 'on' or request.POST.get('is_income') == 'True'
        
        if name:
            try:
                # Restrict to unique per user
                category, created = Category.objects.get_or_create(user=request.user, name=name, is_income=is_income)
                if created:
                    messages.success(request, f"Custom category '{name}' created successfully!")
                else:
                    messages.info(request, f"Category '{name}' already exists.")
            except Exception as e:
                messages.error(request, f"Error creating category: {e}")
        else:
            messages.error(request, "Category name cannot be empty.")
    return redirect('dashboard')


@login_required
def edit_transaction(request, pk):
    tx = get_object_or_404(Transaction, pk=pk, user=request.user)
    if request.method == 'POST':
        amount = request.POST.get('amount')
        tx_type = request.POST.get('transaction_type')
        category = request.POST.get('category')
        tx_date = request.POST.get('date')
        description = request.POST.get('description', '')
        
        try:
            tx.amount = Decimal(amount)
            tx.transaction_type = tx_type
            tx.category = category
            tx.date = tx_date
            tx.description = description
            tx.save()
            messages.success(request, "Transaction updated successfully!")
            
            # Send budget alert emails if transaction is an Expense (OUT)
            if tx_type == 'OUT':
                check_and_send_budget_alerts(request.user, category, tx_date)
        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f"Error updating transaction: {e}")
    return redirect('dashboard')


@login_required
def delete_transaction_bulk(request):
    if request.method == 'POST':
        tx_ids_str = request.POST.get('transaction_ids', '')
        if tx_ids_str:
            try:
                id_list = [int(x) for x in tx_ids_str.split(',') if x.strip()]
                deleted_count, _ = Transaction.objects.filter(user=request.user, id__in=id_list).delete()
                messages.success(request, f"Deleted {deleted_count} selected transaction(s).")
            except Exception as e:
                messages.error(request, f"Error bulk deleting: {e}")
        else:
            messages.warning(request, "No transactions selected for deletion.")
    return redirect('dashboard')


@login_required
def edit_reminder(request, pk):
    reminder = get_object_or_404(Reminder, pk=pk, user=request.user)
    if request.method == 'POST':
        title = request.POST.get('title')
        amount = request.POST.get('amount')
        due_date = request.POST.get('due_date')
        is_recurring = request.POST.get('is_recurring') == 'on'
        recurrence_period = request.POST.get('recurrence_period') if is_recurring else None
        
        try:
            reminder.title = title
            reminder.amount = Decimal(amount)
            reminder.due_date = due_date
            reminder.is_recurring = is_recurring
            reminder.recurrence_period = recurrence_period
            reminder.save()
            messages.success(request, f"Bill reminder '{title}' updated successfully!")
        except Exception as e:
            messages.error(request, f"Error updating reminder: {e}")
    return redirect('dashboard')


@login_required
def delete_reminder(request, pk):
    reminder = get_object_or_404(Reminder, pk=pk, user=request.user)
    title = reminder.title
    reminder.delete()
    messages.success(request, f"Bill reminder '{title}' deleted successfully.")
    return redirect('dashboard')


@login_required
def sync_emails_now(request):
    from django.core.management import call_command
    try:
        call_command('fetch_emails')
        messages.success(request, "Email inbox sync completed successfully!")
    except Exception as e:
        messages.error(request, f"Error synchronizing emails: {e}")
    return redirect('dashboard')


@login_required
def download_template(request):
    from openpyxl import Workbook
    from django.http import HttpResponse
    
    wb = Workbook()
    
    # 1. Setup Transactions Sheet
    ws_tx = wb.active
    ws_tx.title = "Transactions"
    ws_tx.append(["Date (YYYY-MM-DD)", "Amount", "Type (IN/OUT)", "Category", "Description"])
    ws_tx.append(["2026-07-10", 1500.00, "OUT", "Utilities", "Electricity Bill"])
    ws_tx.append(["2026-07-11", 75000.00, "IN", "Salary", "Monthly Salary payment"])
    ws_tx.append(["2026-07-12", 450.50, "OUT", "Food", "Lunch"])
    
    # Adjust column dimensions
    for col in ws_tx.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws_tx.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # 2. Setup Instructions Sheet
    ws_instr = wb.create_sheet(title="Instructions")
    ws_instr.append(["Field Name", "Requirement", "Allowed Values / Examples"])
    ws_instr.append(["Date (YYYY-MM-DD)", "Required date format.", "YYYY-MM-DD (e.g., 2026-07-12)"])
    ws_instr.append(["Amount", "Required positive numeric value.", "Any decimal greater than 0 (e.g., 1250.50)"])
    ws_instr.append(["Type (IN/OUT)", "Transaction direction.", "IN (for Incomes) or OUT (for Expenses)"])
    ws_instr.append(["Category", "The budget category name.", "e.g., Salary, Food, Utilities. If the category does not exist, it will be automatically created for your profile."])
    ws_instr.append(["Description", "Optional narrative detail.", "e.g., Rent payment, Weekly groceries"])
    
    for col in ws_instr.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws_instr.column_dimensions[col_letter].width = max(max_len + 3, 15)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="leoxur_import_template.xlsx"'
    wb.save(response)
    return response


@login_required
def import_transactions(request):
    import csv
    from datetime import datetime, date
    from decimal import Decimal
    from openpyxl import load_workbook
    
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        filename = file.name
        
        success_count = 0
        error_rows = []
        
        try:
            if filename.endswith('.csv'):
                decoded_file = file.read().decode('utf-8').splitlines()
                reader = csv.reader(decoded_file)
                # Skip header
                header = next(reader, None)
                rows = list(reader)
                
                for idx, row in enumerate(rows, start=2):
                    if not row or all(not str(val).strip() for val in row):
                        continue
                    if len(row) < 4:
                        error_rows.append(f"Row {idx}: Missing columns.")
                        continue
                        
                    date_val, amount_val, type_val, category_val = row[0].strip(), row[1].strip(), row[2].strip(), row[3].strip()
                    desc_val = row[4].strip() if len(row) > 4 else ""
                    
                    try:
                        date_parsed = datetime.strptime(date_val, '%Y-%m-%d').date()
                    except Exception:
                        error_rows.append(f"Row {idx}: Invalid date format. Use YYYY-MM-DD.")
                        continue
                        
                    try:
                        amount = Decimal(amount_val)
                        if amount <= 0:
                            raise ValueError()
                    except Exception:
                        error_rows.append(f"Row {idx}: Invalid amount. Must be positive number.")
                        continue
                        
                    tx_type = type_val.upper()
                    if tx_type not in ['IN', 'OUT']:
                        error_rows.append(f"Row {idx}: Invalid type. Use IN or OUT.")
                        continue
                        
                    category_name = category_val
                    if not category_name:
                        error_rows.append(f"Row {idx}: Category cannot be empty.")
                        continue
                        
                    # Auto-create category
                    is_income = (tx_type == 'IN')
                    Category.objects.get_or_create(user=request.user, name=category_name, is_income=is_income)
                    
                    Transaction.objects.create(
                        user=request.user,
                        amount=amount,
                        transaction_type=tx_type,
                        category=category_name,
                        date=date_parsed,
                        description=desc_val
                    )
                    success_count += 1
                    
            elif filename.endswith('.xlsx'):
                wb = load_workbook(file, read_only=True)
                if "Transactions" not in wb.sheetnames:
                    messages.error(request, "Invalid template: Sheet named 'Transactions' was not found.")
                    return redirect('dashboard')
                    
                ws = wb["Transactions"]
                rows = list(ws.iter_rows(values_only=True))
                if not rows or len(rows) <= 1:
                    messages.warning(request, "No transaction data found in the spreadsheet.")
                    return redirect('dashboard')
                    
                # Skip header
                data_rows = rows[1:]
                
                for idx, row in enumerate(data_rows, start=2):
                    if not row or all(val is None for val in row):
                        continue
                    if len(row) < 4:
                        error_rows.append(f"Row {idx}: Missing columns.")
                        continue
                        
                    date_val, amount_val, type_val, category_val = row[0], row[1], row[2], row[3]
                    desc_val = row[4] if len(row) > 4 else ""
                    
                    if isinstance(date_val, datetime):
                        date_parsed = date_val.date()
                    elif isinstance(date_val, date):
                        date_parsed = date_val
                    else:
                        try:
                            date_parsed = datetime.strptime(str(date_val).strip(), '%Y-%m-%d').date()
                        except Exception:
                            error_rows.append(f"Row {idx}: Invalid date format. Use YYYY-MM-DD.")
                            continue
                            
                    try:
                        amount = Decimal(str(amount_val).strip())
                        if amount <= 0:
                            raise ValueError()
                    except Exception:
                        error_rows.append(f"Row {idx}: Invalid amount. Must be positive number.")
                        continue
                        
                    tx_type = str(type_val).strip().upper()
                    if tx_type not in ['IN', 'OUT']:
                        error_rows.append(f"Row {idx}: Invalid type. Use IN or OUT.")
                        continue
                        
                    category_name = str(category_val).strip()
                    if not category_name:
                        error_rows.append(f"Row {idx}: Category cannot be empty.")
                        continue
                        
                    is_income = (tx_type == 'IN')
                    Category.objects.get_or_create(user=request.user, name=category_name, is_income=is_income)
                    
                    Transaction.objects.create(
                        user=request.user,
                        amount=amount,
                        transaction_type=tx_type,
                        category=category_name,
                        date=date_parsed,
                        description=str(desc_val).strip() if desc_val else ""
                    )
                    success_count += 1
            else:
                messages.error(request, "Unsupported file format. Please upload a .csv or .xlsx file.")
                return redirect('dashboard')
                
        except Exception as e:
            messages.error(request, f"Error reading file: {e}")
            return redirect('dashboard')
            
        if error_rows:
            err_msg = " | ".join(error_rows[:5])
            if len(error_rows) > 5:
                err_msg += f" ...and {len(error_rows) - 5} more errors."
            messages.warning(request, f"Imported {success_count} transactions. Ignored rows: {err_msg}")
        else:
            messages.success(request, f"Successfully imported {success_count} transactions!")
            
    else:
        messages.error(request, "No file uploaded.")
        
    return redirect('dashboard')
